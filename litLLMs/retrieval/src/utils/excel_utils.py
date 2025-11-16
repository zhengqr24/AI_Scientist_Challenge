import pandas as pd
import os

from matplotlib.ticker import FormatStrFormatter
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font
from typing import NoReturn
from openpyxl.styles import Alignment


def columns_best_fit(sheet: openpyxl.worksheet.worksheet.Worksheet) -> NoReturn:
    """
    Make all columns best fit
    """
    for column_cells in sheet.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = get_column_letter(column_cells[0].column)
        new_column_length = min(new_column_length, 100)
        if new_column_length > 0:
            sheet.column_dimensions[new_column_letter].width = new_column_length * 1.1
        # Enable text wrapping for the column
        for cell in column_cells:
            cell.alignment = Alignment(wrap_text=True)


def create_spreadsheet(filename, data):
    if not os.path.exists(filename):
        # prepare data for the excel sheet
        wb = Workbook()
        wb.remove(wb.active)  # remove the empty sheet called "Sheet"
        num_sheets = 0
    else:
        wb = load_workbook(filename)
        # get the number of sheets
        num_sheets = len(wb.sheetnames)

    ws = wb.create_sheet(f"Sheet{num_sheets+1}")
    ranking = data["ranking"]
    reason = data["reason"]
    arxiv_ids = data["arxiv_id"]
    reason_verfication = data["ranking_reasoning_verification"]
    query = data["query"]
    paper_ids = data["paper_id"]
    abstracts = data["abstract"]
    s2_query = data.get("s2_keyword_queries", None)
    s2_query_reasoning = data.get("s2_query_reasoning", None)
    s2_query_reasoning_verfication = data.get("s2_query_reasoning_verification", None)
    assert len(paper_ids) == len(abstracts)
    # add Train sentences
    row_num = 1
    ws.append(["Paper ID", "Abstract", "Arxiv ID"])
    ws[f"A{row_num}"].font = Font(bold=True)
    ws[f"B{row_num}"].font = Font(bold=True)
    ws[f"C{row_num}"].font = Font(bold=True)
    # element_ids = np.random.choice(element_ids, 10, replace=False)
    for pid, abs, arxiv_id in zip(paper_ids, abstracts, arxiv_ids):
        ws[f"A{row_num+1}"] = pid
        ws[f"B{row_num+1}"] = abs
        ws[f"C{row_num+1}"] = arxiv_id
        row_num += 1

    # add ranking, reason, and query
    ws["F1"] = "Ranking:"
    ws["F2"] = "Reason:"
    ws["F3"] = "Query:"
    ws["F1"].font = Font(bold=True)
    ws["F2"].font = Font(bold=True)
    ws["F3"].font = Font(bold=True)
    ws["G1"] = ranking
    ws["G2"] = reason
    ws["H2"] = "Extracted Sentence Exists in the abstract?\n\n" + "\n\n".join(
        [f"R{idx}: {str(flag)}" for idx, flag in enumerate(reason_verfication)]
    )
    ws["G3"] = query
    if s2_query is not None:
        ws["F4"] = "Semantic Scholar Query:"
        ws["G4"] = " | ".join(s2_query)
        ws["F4"].font = Font(bold=True)

        ws["F5"] = "Semantic Scholar Query Reasoning:"
        ws["G5"] = "\n\n".join(s2_query_reasoning)
        ws["F5"].font = Font(bold=True)

        ws["H5"] = "Extracted Sentence Exists in the abstract?\n\n" + "\n\n".join(
            [
                f"R{idx}: {str(flag)}"
                for idx, flag in enumerate(s2_query_reasoning_verfication)
            ]
        )

    # adjust column width
    columns_best_fit(ws)

    wb.active = 0
    wb.save(filename)


def save_spreadsheet(
    abstract,
    papers,
    ranking,
    reason,
    arxiv_ids,
    ranking_reasoning_verification,
    extracted_s2_queries,
    s2_query_reasoning,
    query_reasoning_verification,
    output_dir,
    config,
):
    paper_abstracts = {"paper_id": [], "abstract": []}
    for i, paper in enumerate(papers):
        paper_abstracts["paper_id"].append(i + 1)
        title = paper.get("title_paper", "No title")
        a = paper.get("abstract", "Abstract not available")
        paper_abstracts["abstract"].append(f"Title: {title}\nAbstract: {a}")
    data = {
        "query": abstract,
        "ranking": ranking,
        "reason": reason,
        "arxiv_id": arxiv_ids,
        "ranking_reasoning_verification": ranking_reasoning_verification,
    }

    if extracted_s2_queries:
        data.update(
            {
                "s2_keyword_queries": extracted_s2_queries,
                "s2_query_reasoning": s2_query_reasoning,
                "s2_query_reasoning_verification": query_reasoning_verification,
            }
        )
    data.update(paper_abstracts)
    create_spreadsheet(
        os.path.join(
            output_dir,
            f"reasoning_output-{config.gen_engine}-n={config.n_candidates}-n_rec={config.n_queries}-reranking_strategy={config.reranking_prompt_type}.xlsx",
        ),
        data,
    )
    print(f"Saved reasoning output to {output_dir}")
